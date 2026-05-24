"""导出模块 — 将解析后的日志数据导出为 Excel 等格式."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from logcat_viewer.parser import COLUMNS, level_color

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """导出错误。"""


def _check_openpyxl() -> None:
    """检查 openpyxl 是否可用，否则抛出友好的错误信息。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ExportError(
            "缺少 openpyxl 库，无法导出 Excel。\n"
            "请执行: pip install openpyxl"
        )


def export_excel(
    metadata: dict[str, Any],
    entries: list[dict[str, Any]],
    output_path: str | Path,
) -> None:
    """将日志数据导出为带高亮的 Excel 文件。

    Args:
        metadata: 设备元数据。
        entries: 日志条目列表。
        output_path: 输出 .xlsx 文件路径。

    Raises:
        ExportError: openpyxl 未安装或写入失败。
    """
    _check_openpyxl()
    logger.info(f"开始导出 Excel: {output_path}, 共 {len(entries)} 条日志")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output_path = Path(output_path)

    header_fill   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font   = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    thin_border   = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    wrap_align    = Alignment(wrap_text=True, vertical="top")
    data_font     = Font(name="Consolas", size=9)
    even_fill     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    ts_fill   = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    tag_fill  = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
    msg_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    app_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Logcat"

        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_offset, entry in enumerate(entries):
            row = row_offset + 2
            base_fill = even_fill if row_offset % 2 == 0 else odd_fill

            values = [
                entry["index"],
                entry["timestamp"],
                entry["level"],
                entry["pid"],
                entry["tid"],
                entry["application"],
                entry["process"],
                entry["tag"],
                entry["message"],
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = wrap_align
                cell.border = thin_border
                cell.fill = base_fill

                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif col_idx == 2:
                    cell.fill = ts_fill
                elif col_idx == 3:
                    lc = level_color(str(value))
                    if lc:
                        cell.fill = PatternFill(start_color=lc[1:], end_color=lc[1:], fill_type="solid")
                elif col_idx == 6:
                    cell.fill = app_fill
                elif col_idx == 8:
                    cell.fill = tag_fill
                elif col_idx == 9:
                    cell.fill = msg_fill

        col_widths = {1: 6, 2: 24, 3: 9, 4: 8, 5: 8, 6: 30, 7: 30, 8: 22, 9: 100}
        from openpyxl.utils import get_column_letter
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(entries) + 1}"

        ws2 = wb.create_sheet("Device Info")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 80

        info_key_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        info_key_font = Font(name="Consolas", bold=True, size=9)
        info_val_font = Font(name="Consolas", size=9)

        device       = metadata.get("device", {}) or {}
        filter_str   = metadata.get("filter", "")
        project_ids  = metadata.get("projectApplicationIds", [])

        rows_info: list[tuple[str, str]] = [
            ("Device Name",     str(device.get("name", ""))),
            ("Model",           str(device.get("model", ""))),
            ("Serial",          str(device.get("serialNumber", ""))),
            ("Android Release", str(device.get("release", ""))),
            ("SDK",             str(device.get("sdk", ""))),
            ("Feature Level",   str(device.get("featureLevel", ""))),
            ("Type",            str(device.get("type", ""))),
            ("Emulator",        str(device.get("isEmulator", ""))),
            ("Online",          str(device.get("isOnline", ""))),
            ("Device ID",       str(device.get("deviceId", ""))),
            ("", ""),
            ("Log Filter",      str(filter_str)),
            ("", ""),
        ]
        for j, app_id in enumerate(project_ids, start=1):
            rows_info.append((f"Project App #{j}", str(app_id)))

        ws2.insert_rows(1)
        title_cell = ws2.cell(row=1, column=1, value="Device Information")
        title_cell.font = Font(name="Consolas", bold=True, size=12, color="4472C4")

        for i, (key, val) in enumerate(rows_info, start=2):
            ck = ws2.cell(row=i, column=1, value=key)
            cv = ws2.cell(row=i, column=2, value=val)
            ck.font = info_key_font
            ck.fill = info_key_fill
            ck.border = thin_border
            ck.alignment = Alignment(vertical="top")
            cv.font = info_val_font
            cv.border = thin_border
            cv.alignment = Alignment(wrap_text=True, vertical="top")

        try:
            wb.save(str(output_path))
            logger.info(f"Excel 导出成功: {output_path}")
        except OSError as exc:
            logger.error(f"Excel 导出失败: {exc}")
            raise ExportError(f"无法写入文件 ({output_path}): {exc}") from exc
    finally:
        wb.close()
